#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gestion des Équipes - Application de bureau hors ligne
Remplace le classeur Excel "Equipes.xlsm" (horaire hebdomadaire + gestion des employés + heures)

Fonctionne 100% hors ligne. Toutes les données sont sauvegardées automatiquement
dans un fichier data.json situé dans le dossier "GestionEquipes" du dossier
Documents de l'utilisateur, donc rien n'est perdu à la fermeture.
"""

import json
import os
import sys
import datetime
import tkinter as tk
from tkinter import ttk, messagebox

APP_TITLE = "Gestion des Équipes"

DAYS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
QUARTS = ["JOUR", "SOIR"]
TRC_LINES = ["TRC-1", "TRC-2", "TRC-3"]
ROLES = ["Opérateur", "Aide opérateur", "Journalier"]
# (clé interne unique, libellé affiché) - deux entrées peuvent partager le même libellé
AUTRES_POSTES = [
    ("402", "402"),
    ("Imprimerie", "Imprimerie"),
    ("Réception/Expédition", "Réception/Expédition"),
    ("Colle", "Colle"),
    ("Shipping1", "Shipping"),
    ("Shipping2", "Shipping"),
]
AUTRES_POSTES_LABELS = list(dict.fromkeys(label for _, label in AUTRES_POSTES))

# ---------------------------------------------------------------------------
# Emplacement du fichier de données (persiste entre les lancements de l'exe)
# ---------------------------------------------------------------------------
def get_data_dir():
    home = os.path.expanduser("~")
    docs = os.path.join(home, "Documents")
    base = docs if os.path.isdir(docs) else home
    data_dir = os.path.join(base, "GestionEquipes")
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


DATA_FILE = os.path.join(get_data_dir(), "data.json")


def empty_slot():
    return {"nom": "", "heures": ""}


def default_day():
    day = {"quarts": {}, "emballeur": {}, "autres_postes": {}}
    for q in QUARTS:
        day["quarts"][q] = {}
        for trc in TRC_LINES:
            day["quarts"][q][trc] = {role: empty_slot() for role in ROLES}
        day["emballeur"][q] = {"nom": "", "horaire": "", "heures": ""}
    for key, _label in AUTRES_POSTES:
        day["autres_postes"][key] = {"nom": "", "horaire": "", "heures": ""}
    day["postes_divers"] = []  # employés saisonniers: lignes ajoutées librement
    return day


def default_data():
    return {
        "employees": [],  # {id, nom, poste, heures_cible}
        "schedule": {d: default_day() for d in DAYS},
        "next_id": 1,
        "next_divers_id": 1,
    }


def _coerce_slot(value):
    """Upgrade older data formats (plain string) to the {'nom','heures'} dict."""
    if isinstance(value, str):
        return {"nom": value, "heures": ""}
    if isinstance(value, dict):
        value.setdefault("nom", "")
        value.setdefault("heures", "")
        return value
    return empty_slot()


def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            base = default_data()
            for k in base:
                if k not in data:
                    data[k] = base[k]
            for d in DAYS:
                if d not in data["schedule"]:
                    data["schedule"][d] = default_day()
                day = data["schedule"][d]
                for q in QUARTS:
                    day.setdefault("quarts", {}).setdefault(q, {})
                    for trc in TRC_LINES:
                        day["quarts"][q].setdefault(trc, {})
                        for role in ROLES:
                            day["quarts"][q][trc][role] = _coerce_slot(day["quarts"][q][trc].get(role))
                    day.setdefault("emballeur", {}).setdefault(q, {})
                    day["emballeur"][q] = _coerce_slot(day["emballeur"][q])
                    day["emballeur"][q].setdefault("horaire", "")
                day.setdefault("autres_postes", {})
                # migration: l'ancienne clé unique "Shipping" devient "Shipping1"
                if "Shipping" in day["autres_postes"] and "Shipping1" not in day["autres_postes"]:
                    day["autres_postes"]["Shipping1"] = day["autres_postes"].pop("Shipping")
                for key, _label in AUTRES_POSTES:
                    day["autres_postes"][key] = _coerce_slot(day["autres_postes"].get(key))
                    day["autres_postes"][key].setdefault("horaire", "")
                day.setdefault("postes_divers", [])
                for entry in day["postes_divers"]:
                    entry.setdefault("nom", "")
                    entry.setdefault("ligne", "")
                    entry.setdefault("horaire", "")
                    entry.setdefault("heures", "")
            data.setdefault("next_divers_id", 1)
            # drop obsolete "heures" dict from very first version, no longer used
            data.pop("heures", None)
            return data
        except Exception:
            messagebox.showwarning(
                APP_TITLE,
                "Le fichier de données existant est corrompu.\n"
                "Une nouvelle base de données vide a été créée.\n"
                f"(ancien fichier conservé à: {DATA_FILE}.bak)",
            )
            try:
                os.replace(DATA_FILE, DATA_FILE + ".bak")
            except Exception:
                pass
    return default_data()


def to_float(text):
    try:
        return float(str(text).replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def compute_weekly_hours(data):
    """Somme des heures travaillées entrées dans TOUS les jours, par employé."""
    totals = {e["nom"]: 0.0 for e in data["employees"]}
    for d in DAYS:
        day = data["schedule"][d]
        for q in QUARTS:
            for trc in TRC_LINES:
                for role in ROLES:
                    slot = day["quarts"][q][trc][role]
                    nom = slot.get("nom", "")
                    if nom:
                        totals[nom] = totals.get(nom, 0.0) + to_float(slot.get("heures"))
            emb = day["emballeur"][q]
            if emb.get("nom"):
                totals[emb["nom"]] = totals.get(emb["nom"], 0.0) + to_float(emb.get("heures"))
        for key, _label in AUTRES_POSTES:
            info = day["autres_postes"][key]
            if info.get("nom"):
                totals[info["nom"]] = totals.get(info["nom"], 0.0) + to_float(info.get("heures"))
        for entry in day.get("postes_divers", []):
            if entry.get("nom"):
                totals[entry["nom"]] = totals.get(entry["nom"], 0.0) + to_float(entry.get("heures"))
    return totals


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1400x780")
        self.minsize(1100, 650)

        self.data = load_data()

        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("Treeview", rowheight=26, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.configure("TNotebook.Tab", font=("Segoe UI", 10, "bold"), padding=[12, 6])

        self._build_menu()

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=6, pady=6)

        self.employees_tab = EmployeesTab(self.notebook, self)
        self.notebook.add(self.employees_tab, text="Employés")

        self.day_tabs = {}
        for d in DAYS:
            tab = DayTab(self.notebook, self, d)
            self.day_tabs[d] = tab
            self.notebook.add(tab, text=d)

        self.heures_tab = HeuresTab(self.notebook, self)
        self.notebook.add(self.heures_tab, text="Heures (semaine)")

        self.status = tk.StringVar(value=f"Données sauvegardées dans: {DATA_FILE}")
        status_bar = ttk.Label(self, textvariable=self.status, anchor="w",
                                relief="sunken", padding=(6, 2))
        status_bar.pack(fill="x", side="bottom")

        self.protocol("WM_DELETE_WINDOW", self.on_close)

    # ------------------------------------------------------------------
    def _build_menu(self):
        menubar = tk.Menu(self)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Sauvegarder maintenant", command=self.save)
        filemenu.add_command(label="Exporter en Excel (.xlsx)...", command=self.export_excel)
        filemenu.add_separator()
        filemenu.add_command(label="Quitter", command=self.on_close)
        menubar.add_cascade(label="Fichier", menu=filemenu)

        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="À propos", command=self.show_about)
        menubar.add_cascade(label="Aide", menu=helpmenu)
        self.config(menu=menubar)

    def show_about(self):
        messagebox.showinfo(
            APP_TITLE,
            f"{APP_TITLE}\n\n"
            "Application hors ligne pour gérer l'horaire hebdomadaire, "
            "les employés et les heures travaillées.\n\n"
            f"Fichier de données: {DATA_FILE}",
        )

    # ------------------------------------------------------------------
    def employee_names(self):
        return [e["nom"] for e in self.data["employees"]]

    def refresh_all_employee_dropdowns(self):
        names = [""] + sorted(self.employee_names())
        for tab in self.day_tabs.values():
            tab.update_employee_list(names)
        self.refresh_hours_everywhere()

    def refresh_hours_everywhere(self):
        """Recalcule le cumul d'heures et met à jour tous les panneaux qui l'affichent."""
        for tab in self.day_tabs.values():
            tab.refresh_side_panel()
        self.heures_tab.refresh()

    def save(self, silent=False):
        try:
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
            self.status.set(
                f"Sauvegardé à {datetime.datetime.now().strftime('%H:%M:%S')} — {DATA_FILE}"
            )
        except Exception as e:
            if not silent:
                messagebox.showerror(APP_TITLE, f"Erreur de sauvegarde:\n{e}")

    def on_close(self):
        self.save(silent=True)
        self.destroy()

    # ------------------------------------------------------------------
    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            messagebox.showerror(
                APP_TITLE,
                "L'export Excel nécessite le module 'openpyxl'.\n"
                "S'il n'est pas disponible, réinstallez l'application avec "
                "requirements.txt (voir README).",
            )
            return

        from tkinter import filedialog

        path = filedialog.asksaveasfilename(
            title="Exporter en Excel",
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Horaire_Equipes.xlsx",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="D9E1F2")
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for d in DAYS:
            ws = wb.create_sheet(d)
            day = self.data["schedule"][d]
            row = 1
            ws.cell(row=row, column=1, value=d.upper()).font = Font(bold=True, size=14)
            row += 2
            for q in QUARTS:
                ws.cell(row=row, column=1, value=q).font = bold
                row += 1
                headers = ["Poste"] + TRC_LINES
                for c, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=c, value=h)
                    cell.font = bold
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center
                row += 1
                for role in ROLES:
                    ws.cell(row=row, column=1, value=role).border = border
                    for c, trc in enumerate(TRC_LINES, start=2):
                        slot = day["quarts"][q][trc][role]
                        val = slot.get("nom", "")
                        h = slot.get("heures", "")
                        cell = ws.cell(row=row, column=c, value=f"{val} ({h}h)" if val and h else val)
                        cell.border = border
                    row += 1
                emb = day["emballeur"][q]
                ws.cell(row=row, column=1, value="Emballeur").border = border
                ws.cell(row=row, column=2, value=emb.get("nom", "")).border = border
                ws.cell(row=row, column=3, value=emb.get("horaire", "")).border = border
                ws.cell(row=row, column=4, value=emb.get("heures", "")).border = border
                row += 2

            ws.cell(row=row, column=1, value="Autres postes").font = bold
            row += 1
            for key, label in AUTRES_POSTES:
                info = day["autres_postes"][key]
                ws.cell(row=row, column=1, value=label).border = border
                ws.cell(row=row, column=2, value=info.get("nom", "")).border = border
                ws.cell(row=row, column=3, value=info.get("horaire", "")).border = border
                ws.cell(row=row, column=4, value=info.get("heures", "")).border = border
                row += 1

            if day.get("postes_divers"):
                row += 1
                ws.cell(row=row, column=1, value="Postes divers (saisonniers)").font = bold
                row += 1
                headers = ["Nom", "Ligne de production", "Horaire", "Heures"]
                for c, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=c, value=h)
                    cell.font = bold
                    cell.fill = header_fill
                    cell.border = border
                row += 1
                for entry in day["postes_divers"]:
                    ws.cell(row=row, column=1, value=entry.get("nom", "")).border = border
                    ws.cell(row=row, column=2, value=entry.get("ligne", "")).border = border
                    ws.cell(row=row, column=3, value=entry.get("horaire", "")).border = border
                    ws.cell(row=row, column=4, value=entry.get("heures", "")).border = border
                    row += 1

            for col, width in zip("ABCD", (22, 22, 22, 14)):
                ws.column_dimensions[col].width = width

        ws = wb.create_sheet("Employés")
        headers = ["Nom", "Poste par défaut", "Heures cible / semaine"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        for r, emp in enumerate(self.data["employees"], start=2):
            ws.cell(row=r, column=1, value=emp["nom"])
            ws.cell(row=r, column=2, value=emp.get("poste", ""))
            ws.cell(row=r, column=3, value=emp.get("heures_cible", 40))
        for col, width in zip("ABC", (28, 22, 20)):
            ws.column_dimensions[col].width = width

        ws = wb.create_sheet("Heures")
        headers = ["Nom", "Heures cible", "Heures cumulées", "Écart"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
        totals = compute_weekly_hours(self.data)
        for r, emp in enumerate(self.data["employees"], start=2):
            nom = emp["nom"]
            cible = emp.get("heures_cible", 40)
            ws.cell(row=r, column=1, value=nom)
            ws.cell(row=r, column=2, value=cible)
            ws.cell(row=r, column=3, value=totals.get(nom, 0.0))
            ws.cell(row=r, column=4, value=f"=C{r}-B{r}")
        for col, width in zip("ABCD", (28, 14, 16, 10)):
            ws.column_dimensions[col].width = width

        try:
            wb.save(path)
            messagebox.showinfo(APP_TITLE, f"Export réussi:\n{path}")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Erreur lors de l'export:\n{e}")


# ===========================================================================
class EmployeesTab(ttk.Frame):
    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app
        self._build()
        self.refresh()

    def _build(self):
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Nom complet:").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.nom_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.nom_var, width=30).grid(row=0, column=1, padx=4, pady=4)

        ttk.Label(top, text="Poste par défaut:").grid(row=0, column=2, sticky="w", padx=4, pady=4)
        self.poste_var = tk.StringVar()
        poste_cb = ttk.Combobox(
            top, textvariable=self.poste_var, width=22,
            values=ROLES + ["Emballeur"] + AUTRES_POSTES_LABELS,
        )
        poste_cb.grid(row=0, column=3, padx=4, pady=4)

        ttk.Label(top, text="Heures cible / semaine:").grid(row=0, column=4, sticky="w", padx=4, pady=4)
        self.heures_var = tk.StringVar(value="40")
        ttk.Spinbox(top, from_=0, to=80, textvariable=self.heures_var, width=8).grid(
            row=0, column=5, padx=4, pady=4
        )

        btns = ttk.Frame(top)
        btns.grid(row=0, column=6, padx=10)
        ttk.Button(btns, text="➕ Ajouter", command=self.add_employee).pack(side="left", padx=2)
        ttk.Button(btns, text="✏ Modifier", command=self.edit_employee).pack(side="left", padx=2)
        ttk.Button(btns, text="🗑 Supprimer", command=self.remove_employee).pack(side="left", padx=2)

        cols = ("nom", "poste", "heures")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", selectmode="browse")
        self.tree.heading("nom", text="Nom")
        self.tree.heading("poste", text="Poste par défaut")
        self.tree.heading("heures", text="Heures cible / semaine")
        self.tree.column("nom", width=300)
        self.tree.column("poste", width=220)
        self.tree.column("heures", width=180, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Double-1>", lambda e: self.edit_employee())

        info = ttk.Label(
            self,
            text=("Astuce: ajoutez ici tous vos employés une seule fois. Ils apparaîtront "
                  "ensuite dans les listes déroulantes de chaque jour et dans l'onglet Heures."),
            foreground="#555",
        )
        info.pack(fill="x", padx=10, pady=(0, 8))

    def on_select(self, _event=None):
        sel = self.tree.selection()
        if not sel:
            return
        emp = self._employee_by_iid(sel[0])
        if emp:
            self.nom_var.set(emp["nom"])
            self.poste_var.set(emp.get("poste", ""))
            self.heures_var.set(str(emp.get("heures_cible", 40)))

    def _employee_by_iid(self, iid):
        for e in self.app.data["employees"]:
            if str(e["id"]) == iid:
                return e
        return None

    def add_employee(self):
        nom = self.nom_var.get().strip()
        if not nom:
            messagebox.showwarning(APP_TITLE, "Veuillez entrer un nom.")
            return
        if nom.upper() in [e["nom"].upper() for e in self.app.data["employees"]]:
            messagebox.showwarning(APP_TITLE, "Un employé avec ce nom existe déjà.")
            return
        try:
            heures = float(self.heures_var.get())
        except ValueError:
            heures = 40
        emp = {
            "id": self.app.data["next_id"],
            "nom": nom,
            "poste": self.poste_var.get().strip(),
            "heures_cible": heures,
        }
        self.app.data["next_id"] += 1
        self.app.data["employees"].append(emp)
        self.app.save()
        self.refresh()
        self.app.refresh_all_employee_dropdowns()
        self.nom_var.set("")
        self.poste_var.set("")
        self.heures_var.set("40")

    def edit_employee(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "Sélectionnez d'abord un employé dans la liste.")
            return
        emp = self._employee_by_iid(sel[0])
        if not emp:
            return
        old_nom = emp["nom"]
        nom = self.nom_var.get().strip()
        if not nom:
            messagebox.showwarning(APP_TITLE, "Le nom ne peut pas être vide.")
            return
        try:
            heures = float(self.heures_var.get())
        except ValueError:
            heures = 40
        emp["nom"] = nom
        emp["poste"] = self.poste_var.get().strip()
        emp["heures_cible"] = heures

        if old_nom != nom:
            self._rename_in_schedule(old_nom, nom)

        self.app.save()
        self.refresh()
        self.app.refresh_all_employee_dropdowns()

    def _rename_in_schedule(self, old, new):
        for d in DAYS:
            day = self.app.data["schedule"][d]
            for q in QUARTS:
                for trc in TRC_LINES:
                    for role in ROLES:
                        if day["quarts"][q][trc][role].get("nom") == old:
                            day["quarts"][q][trc][role]["nom"] = new
                if day["emballeur"][q].get("nom") == old:
                    day["emballeur"][q]["nom"] = new
            for key, _label in AUTRES_POSTES:
                if day["autres_postes"][key].get("nom") == old:
                    day["autres_postes"][key]["nom"] = new
            for entry in day.get("postes_divers", []):
                if entry.get("nom") == old:
                    entry["nom"] = new

    def _clear_in_schedule(self, name):
        for d in DAYS:
            day = self.app.data["schedule"][d]
            for q in QUARTS:
                for trc in TRC_LINES:
                    for role in ROLES:
                        if day["quarts"][q][trc][role].get("nom") == name:
                            day["quarts"][q][trc][role] = empty_slot()
                if day["emballeur"][q].get("nom") == name:
                    day["emballeur"][q]["nom"] = ""
                    day["emballeur"][q]["heures"] = ""
            for key, _label in AUTRES_POSTES:
                if day["autres_postes"][key].get("nom") == name:
                    day["autres_postes"][key]["nom"] = ""
                    day["autres_postes"][key]["heures"] = ""
            day["postes_divers"] = [
                e for e in day.get("postes_divers", []) if e.get("nom") != name
            ]

    def remove_employee(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo(APP_TITLE, "Sélectionnez d'abord un employé dans la liste.")
            return
        emp = self._employee_by_iid(sel[0])
        if not emp:
            return
        if not messagebox.askyesno(
            APP_TITLE,
            f"Supprimer définitivement '{emp['nom']}' ?\n"
            "Il sera aussi retiré de tous les horaires où il était assigné.",
        ):
            return
        self._clear_in_schedule(emp["nom"])
        self.app.data["employees"] = [
            e for e in self.app.data["employees"] if e["id"] != emp["id"]
        ]
        self.app.save()
        self.refresh()
        self.app.refresh_all_employee_dropdowns()
        for tab in self.app.day_tabs.values():
            tab.rebuild()
        self.nom_var.set("")
        self.poste_var.set("")
        self.heures_var.set("40")

    def refresh(self):
        self.tree.delete(*self.tree.get_children())
        for e in sorted(self.app.data["employees"], key=lambda x: x["nom"]):
            self.tree.insert(
                "", "end", iid=str(e["id"]),
                values=(e["nom"], e.get("poste", ""), e.get("heures_cible", 40)),
            )


# ===========================================================================
class DayTab(ttk.Frame):
    """One tab per weekday reproducing the shift grid, with hours-per-slot inputs
    and a live side panel showing cumulative / remaining hours for every employee."""

    def __init__(self, parent, app: App, day_name: str):
        super().__init__(parent)
        self.app = app
        self.day_name = day_name
        self.combos = []  # comboboxes to refresh with the employee list
        self.side_tree = None
        self._build()

    def rebuild(self):
        for child in self.winfo_children():
            child.destroy()
        self.combos = []
        self._build()

    def _build(self):
        # Horizontal split: schedule (scrollable, left) | side panel (right, fixed)
        left_container = ttk.Frame(self)
        left_container.pack(side="left", fill="both", expand=True)

        right_container = ttk.Frame(self, width=300)
        right_container.pack(side="right", fill="y")
        right_container.pack_propagate(False)

        canvas = tk.Canvas(left_container, highlightthickness=0)
        vsb = ttk.Scrollbar(left_container, orient="vertical", command=canvas.yview)
        container = ttk.Frame(canvas)
        container.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=container, anchor="nw")
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        day = self.app.data["schedule"][self.day_name]

        for q in QUARTS:
            frame = ttk.LabelFrame(container, text=f"Quart de {q}", padding=10)
            frame.pack(fill="x", padx=10, pady=8)

            for col, trc in enumerate(TRC_LINES):
                block = ttk.Frame(frame, padding=6, relief="groove", borderwidth=1)
                block.grid(row=0, column=col, padx=6, pady=4, sticky="n")
                ttk.Label(block, text=trc, font=("Segoe UI", 10, "bold")).pack(anchor="w")
                for role in ROLES:
                    row = ttk.Frame(block)
                    row.pack(fill="x", pady=2)
                    ttk.Label(row, text=role + ":", width=14).pack(side="left")
                    slot = day["quarts"][q][trc][role]

                    nvar = tk.StringVar(value=slot.get("nom", ""))
                    cb = ttk.Combobox(row, textvariable=nvar, width=18,
                                       values=[""] + sorted(self.app.employee_names()))
                    cb.pack(side="left")
                    cb.bind("<<ComboboxSelected>>",
                            lambda e, q=q, trc=trc, role=role, var=nvar: self._set_slot_field(q, trc, role, "nom", var.get()))
                    cb.bind("<FocusOut>",
                            lambda e, q=q, trc=trc, role=role, var=nvar: self._set_slot_field(q, trc, role, "nom", var.get()))
                    self.combos.append(cb)

                    ttk.Label(row, text="h:").pack(side="left", padx=(6, 2))
                    hvar = tk.StringVar(value=slot.get("heures", ""))
                    hent = ttk.Entry(row, textvariable=hvar, width=5)
                    hent.pack(side="left")
                    hent.bind("<FocusOut>",
                              lambda e, q=q, trc=trc, role=role, var=hvar: self._set_slot_field(q, trc, role, "heures", var.get()))
                    hent.bind("<Return>",
                              lambda e, q=q, trc=trc, role=role, var=hvar: self._set_slot_field(q, trc, role, "heures", var.get()))

            emb_block = ttk.Frame(frame, padding=6, relief="groove", borderwidth=1)
            emb_block.grid(row=0, column=3, padx=6, pady=4, sticky="n")
            ttk.Label(emb_block, text="Emballeur", font=("Segoe UI", 10, "bold")).pack(anchor="w")

            row = ttk.Frame(emb_block)
            row.pack(fill="x", pady=2)
            ttk.Label(row, text="Nom:", width=8).pack(side="left")
            emb = day["emballeur"][q]
            nv = tk.StringVar(value=emb.get("nom", ""))
            cb = ttk.Combobox(row, textvariable=nv, width=16, values=[""] + sorted(self.app.employee_names()))
            cb.pack(side="left")
            cb.bind("<<ComboboxSelected>>", lambda e, q=q, var=nv: self._set_emballeur(q, "nom", var.get()))
            cb.bind("<FocusOut>", lambda e, q=q, var=nv: self._set_emballeur(q, "nom", var.get()))
            self.combos.append(cb)

            row2 = ttk.Frame(emb_block)
            row2.pack(fill="x", pady=2)
            ttk.Label(row2, text="Horaire:", width=8).pack(side="left")
            hv = tk.StringVar(value=emb.get("horaire", ""))
            ent = ttk.Entry(row2, textvariable=hv, width=16)
            ent.pack(side="left")
            ent.bind("<FocusOut>", lambda e, q=q, var=hv: self._set_emballeur(q, "horaire", var.get()))
            ent.bind("<Return>", lambda e, q=q, var=hv: self._set_emballeur(q, "horaire", var.get()))

            row3 = ttk.Frame(emb_block)
            row3.pack(fill="x", pady=2)
            ttk.Label(row3, text="Heures:", width=8).pack(side="left")
            hrv = tk.StringVar(value=emb.get("heures", ""))
            hrent = ttk.Entry(row3, textvariable=hrv, width=6)
            hrent.pack(side="left")
            hrent.bind("<FocusOut>", lambda e, q=q, var=hrv: self._set_emballeur(q, "heures", var.get()))
            hrent.bind("<Return>", lambda e, q=q, var=hrv: self._set_emballeur(q, "heures", var.get()))

        autres_frame = ttk.LabelFrame(container, text="Autres postes", padding=10)
        autres_frame.pack(fill="x", padx=10, pady=8)
        for i, (poste_key, poste_label) in enumerate(AUTRES_POSTES):
            block = ttk.Frame(autres_frame, padding=6, relief="groove", borderwidth=1)
            block.grid(row=0, column=i, padx=6, pady=4, sticky="n")
            ttk.Label(block, text=poste_label, font=("Segoe UI", 10, "bold")).pack(anchor="w")
            info = day["autres_postes"][poste_key]

            row = ttk.Frame(block)
            row.pack(fill="x", pady=2)
            ttk.Label(row, text="Nom:", width=6).pack(side="left")
            nv = tk.StringVar(value=info.get("nom", ""))
            cb = ttk.Combobox(row, textvariable=nv, width=16, values=[""] + sorted(self.app.employee_names()))
            cb.pack(side="left")
            cb.bind("<<ComboboxSelected>>", lambda e, k=poste_key, var=nv: self._set_autre(k, "nom", var.get()))
            cb.bind("<FocusOut>", lambda e, k=poste_key, var=nv: self._set_autre(k, "nom", var.get()))
            self.combos.append(cb)

            row2 = ttk.Frame(block)
            row2.pack(fill="x", pady=2)
            ttk.Label(row2, text="Horaire:", width=6).pack(side="left")
            hv = tk.StringVar(value=info.get("horaire", ""))
            ent = ttk.Entry(row2, textvariable=hv, width=16)
            ent.pack(side="left")
            ent.bind("<FocusOut>", lambda e, k=poste_key, var=hv: self._set_autre(k, "horaire", var.get()))
            ent.bind("<Return>", lambda e, k=poste_key, var=hv: self._set_autre(k, "horaire", var.get()))

            row3 = ttk.Frame(block)
            row3.pack(fill="x", pady=2)
            ttk.Label(row3, text="Heures:", width=6).pack(side="left")
            hrv = tk.StringVar(value=info.get("heures", ""))
            hrent = ttk.Entry(row3, textvariable=hrv, width=6)
            hrent.pack(side="left")
            hrent.bind("<FocusOut>", lambda e, k=poste_key, var=hrv: self._set_autre(k, "heures", var.get()))
            hrent.bind("<Return>", lambda e, k=poste_key, var=hrv: self._set_autre(k, "heures", var.get()))

        # ---- Postes divers (employés saisonniers, lignes libres) ------
        divers_frame = ttk.LabelFrame(
            container, text="Postes divers (employés saisonniers)", padding=10
        )
        divers_frame.pack(fill="x", padx=10, pady=8)
        ttk.Label(
            divers_frame,
            text=("Ajoutez ici les employés saisonniers sur des postes variables. "
                  "Autant de lignes que nécessaire."),
            foreground="#555", wraplength=1000, justify="left",
        ).pack(anchor="w", pady=(0, 6))
        ttk.Button(
            divers_frame, text="➕ Ajouter une ligne", command=self._add_divers_row
        ).pack(anchor="w", pady=(0, 8))

        header = ttk.Frame(divers_frame)
        header.pack(fill="x")
        for text, w in (("Nom", 20), ("Ligne de production", 22), ("Horaire", 16), ("Heures", 8)):
            ttk.Label(header, text=text, width=w, font=("Segoe UI", 9, "bold")).pack(side="left", padx=2)

        self.divers_rows_frame = ttk.Frame(divers_frame)
        self.divers_rows_frame.pack(fill="x")
        for entry in day["postes_divers"]:
            self._build_divers_row(entry)

        # ---- Side panel: live employee hours summary -----------------
        ttk.Label(right_container, text="Heures — cumul semaine", font=("Segoe UI", 11, "bold")).pack(
            anchor="w", padx=8, pady=(8, 2)
        )
        ttk.Label(
            right_container,
            text="Se met à jour automatiquement quand vous entrez des heures\n"
                 "(dans ce jour ou dans les autres jours).",
            foreground="#666", justify="left", wraplength=280,
        ).pack(anchor="w", padx=8, pady=(0, 6))

        cols = ("nom", "cumul", "restant")
        self.side_tree = ttk.Treeview(right_container, columns=cols, show="headings", height=20)
        self.side_tree.heading("nom", text="Employé")
        self.side_tree.heading("cumul", text="Fait")
        self.side_tree.heading("restant", text="Reste")
        self.side_tree.column("nom", width=150)
        self.side_tree.column("cumul", width=55, anchor="center")
        self.side_tree.column("restant", width=55, anchor="center")
        self.side_tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self.side_tree.tag_configure("neg", foreground="#B00020")
        self.side_tree.tag_configure("pos", foreground="#1B5E20")

        self.refresh_side_panel()

    # ------------------------------------------------------------------
    def _set_slot_field(self, q, trc, role, field, value):
        self.app.data["schedule"][self.day_name]["quarts"][q][trc][role][field] = value
        self.app.save(silent=True)
        if field == "heures":
            self.app.refresh_hours_everywhere()

    def _set_emballeur(self, q, field, value):
        self.app.data["schedule"][self.day_name]["emballeur"][q][field] = value
        self.app.save(silent=True)
        if field == "heures":
            self.app.refresh_hours_everywhere()

    def _set_autre(self, poste, field, value):
        self.app.data["schedule"][self.day_name]["autres_postes"][poste][field] = value
        self.app.save(silent=True)
        if field == "heures":
            self.app.refresh_hours_everywhere()

    def _build_divers_row(self, entry):
        row = ttk.Frame(self.divers_rows_frame)
        row.pack(fill="x", pady=2)
        entry_id = entry["id"]

        nv = tk.StringVar(value=entry.get("nom", ""))
        cb = ttk.Combobox(row, textvariable=nv, width=18, values=[""] + sorted(self.app.employee_names()))
        cb.pack(side="left", padx=2)
        cb.bind("<<ComboboxSelected>>", lambda e, i=entry_id, var=nv: self._set_divers_field(i, "nom", var.get()))
        cb.bind("<FocusOut>", lambda e, i=entry_id, var=nv: self._set_divers_field(i, "nom", var.get()))
        self.combos.append(cb)

        lv = tk.StringVar(value=entry.get("ligne", ""))
        lent = ttk.Entry(row, textvariable=lv, width=22)
        lent.pack(side="left", padx=2)
        lent.bind("<FocusOut>", lambda e, i=entry_id, var=lv: self._set_divers_field(i, "ligne", var.get()))
        lent.bind("<Return>", lambda e, i=entry_id, var=lv: self._set_divers_field(i, "ligne", var.get()))

        hv = tk.StringVar(value=entry.get("horaire", ""))
        hent = ttk.Entry(row, textvariable=hv, width=16)
        hent.pack(side="left", padx=2)
        hent.bind("<FocusOut>", lambda e, i=entry_id, var=hv: self._set_divers_field(i, "horaire", var.get()))
        hent.bind("<Return>", lambda e, i=entry_id, var=hv: self._set_divers_field(i, "horaire", var.get()))

        hrv = tk.StringVar(value=entry.get("heures", ""))
        hrent = ttk.Entry(row, textvariable=hrv, width=8)
        hrent.pack(side="left", padx=2)
        hrent.bind("<FocusOut>", lambda e, i=entry_id, var=hrv: self._set_divers_field(i, "heures", var.get()))
        hrent.bind("<Return>", lambda e, i=entry_id, var=hrv: self._set_divers_field(i, "heures", var.get()))

        ttk.Button(row, text="🗑", width=3, command=lambda i=entry_id: self._delete_divers_row(i)).pack(
            side="left", padx=4
        )

    def _add_divers_row(self):
        new_id = self.app.data.get("next_divers_id", 1)
        self.app.data["next_divers_id"] = new_id + 1
        entry = {"id": new_id, "nom": "", "ligne": "", "horaire": "", "heures": ""}
        self.app.data["schedule"][self.day_name]["postes_divers"].append(entry)
        self.app.save()
        self._build_divers_row(entry)

    def _delete_divers_row(self, entry_id):
        lst = self.app.data["schedule"][self.day_name]["postes_divers"]
        lst[:] = [e for e in lst if e["id"] != entry_id]
        self.app.save()
        self.app.refresh_hours_everywhere()
        self.rebuild()

    def _set_divers_field(self, entry_id, field, value):
        for e in self.app.data["schedule"][self.day_name]["postes_divers"]:
            if e["id"] == entry_id:
                e[field] = value
                break
        self.app.save(silent=True)
        if field == "heures":
            self.app.refresh_hours_everywhere()

    def update_employee_list(self, names):
        for cb in self.combos:
            current = cb.get()
            cb.configure(values=names)
            cb.set(current)

    def refresh_side_panel(self):
        if self.side_tree is None:
            return
        self.side_tree.delete(*self.side_tree.get_children())
        totals = compute_weekly_hours(self.app.data)
        for emp in sorted(self.app.data["employees"], key=lambda x: x["nom"]):
            nom = emp["nom"]
            cible = emp.get("heures_cible", 40)
            fait = totals.get(nom, 0.0)
            reste = cible - fait
            tag = "neg" if reste < 0 else ("pos" if reste == 0 else "")
            self.side_tree.insert("", "end", values=(nom, f"{fait:g}", f"{reste:g}"), tags=(tag,))


# ===========================================================================
class HeuresTab(ttk.Frame):
    """Résumé hebdomadaire: heures cible vs heures cumulées (calculées
    automatiquement à partir des heures entrées dans chaque jour) et l'écart."""

    def __init__(self, parent, app: App):
        super().__init__(parent)
        self.app = app
        self._build()
        self.refresh()

    def _build(self):
        info = ttk.Label(
            self,
            text=("Les heures réelles sont calculées automatiquement à partir des heures "
                  "entrées sur chaque jour (Lundi à Vendredi). Cet onglet se met à jour tout seul."),
            foreground="#555", padding=10,
        )
        info.pack(fill="x")

        cols = ("nom", "cible", "reel", "ecart")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", selectmode="browse")
        self.tree.heading("nom", text="Nom")
        self.tree.heading("cible", text="Heures cible")
        self.tree.heading("reel", text="Heures cumulées")
        self.tree.heading("ecart", text="Écart")
        self.tree.column("nom", width=300)
        self.tree.column("cible", width=150, anchor="center")
        self.tree.column("reel", width=170, anchor="center")
        self.tree.column("ecart", width=120, anchor="center")
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        total_frame = ttk.Frame(self, padding=10)
        total_frame.pack(fill="x")
        self.total_var = tk.StringVar()
        ttk.Label(total_frame, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(anchor="w")

    def refresh(self):
        self.tree.delete(*self.tree.get_children())
        totals = compute_weekly_hours(self.app.data)
        total_ecart = 0.0
        for emp in sorted(self.app.data["employees"], key=lambda x: x["nom"]):
            nom = emp["nom"]
            cible = emp.get("heures_cible", 40)
            reel = totals.get(nom, 0.0)
            ecart = reel - cible
            total_ecart += ecart
            tag = "neg" if ecart < 0 else ("pos" if ecart > 0 else "")
            self.tree.insert("", "end", values=(nom, f"{cible:g}", f"{reel:g}", f"{ecart:+g}"), tags=(tag,))
        self.tree.tag_configure("neg", foreground="#B00020")
        self.tree.tag_configure("pos", foreground="#1B5E20")
        self.total_var.set(f"Écart total de l'équipe: {total_ecart:+.1f} h")


if __name__ == "__main__":
    app = App()
    app.mainloop()
